from openpyxl import Workbook, load_workbook
from app.models import Order, Sample, User
import json
from django.utils import timezone

def export_to_excel():
    workbook = Workbook()
    
    # Write data to the "Orders" sheet
    orders_sheet = workbook.active
    orders_sheet.title = "Orders"

    # Write headers for the "Orders" sheet
    orders_headers = ['User', 'Name', 'Billing Address', 'AG and HZI', 'Date', 'Quote No', 'Contact Phone', 'Email',
                      'Data Delivery', 'Signature', 'Experiment Title', 'DNA', 'RNA', 'Library', 'Method', 'Buffer',
                      'Organism', 'Isolated From', 'Isolation Method']
    orders_sheet.append(orders_headers)

    # Write data from Order model
    orders = Order.objects.all()
    for order in orders:
        row = [order.user.username, order.name, order.billing_address, order.ag_and_hzi, order.date, order.quote_no,
               str(order.contact_phone), order.email, order.data_delivery, order.signature, order.experiment_title,
               order.dna, order.rna, order.library, order.method, order.buffer, order.organism, order.isolated_from,
               order.isolation_method]
        orders_sheet.append(row)

    # Write data to the "Samples" sheet
    samples_sheet = workbook.create_sheet(title="Samples")

    # Write headers for the "Samples" sheet
    samples_headers = ['Order User', 'Sample Name', 'Concentration', 'Volume', 'Ratio 260/280', 'Ratio 260/230',
                       'Comments', 'Internal ID', 'MIxS Metadata Standard', 'MIxS Metadata']
    samples_sheet.append(samples_headers)

    # Write data from Sample model
    samples = Sample.objects.all()
    for sample in samples:
        internal_id = sample.internal_id.replace(tzinfo=None)  # Remove timezone information
        row = [sample.order.user.username, sample.sample_id, sample.concentration, sample.volume,
               sample.ratio_260_280, sample.ratio_260_230, sample.comments, internal_id,
               sample.mixs_metadata_standard, json.dumps(sample.mixs_metadata)]
        samples_sheet.append(row)

    workbook.save('output.xlsx')

def import_from_excel(file_path):
    workbook = load_workbook(file_path)
    
    # Read data from the "Orders" sheet
    orders_sheet = workbook["Orders"]
    orders_rows = list(orders_sheet.iter_rows(min_row=2, values_only=True))

    for row in orders_rows:
        user = User.objects.get(username=row[0])
        order = Order(user=user, name=row[1], billing_address=row[2], ag_and_hzi=row[3], date=row[4],
                      quote_no=row[5], contact_phone=row[6], email=row[7], data_delivery=row[8],
                      signature=row[9], experiment_title=row[10], dna=row[11], rna=row[12], library=row[13],
                      method=row[14], buffer=row[15], organism=row[16], isolated_from=row[17],
                      isolation_method=row[18])
        order.save()

    samples_sheet = workbook["Samples"]
    samples_rows = list(samples_sheet.iter_rows(min_row=2, values_only=True))

    for row in samples_rows:
        orders = Order.objects.filter(user__username=row[0])
        for order in orders:
            sample = Sample(order=order, sample_id=row[1], concentration=row[2], volume=row[3],
                            ratio_260_280=row[4], ratio_260_230=row[5], comments=row[6],
                            internal_id=timezone.make_aware(row[7]),
                            mixs_metadata_standard=row[8], mixs_metadata=json.loads(row[9]))
            sample.save()